import time
from openpyxl import load_workbook, Workbook
from datetime import datetime
from descricao_carga import getInputs

# Variável global para controlar a primeira consulta
primeira_consulta = True

# Iniciar a cronometragem
start_time = None

# Tentar carregar o workbook existente, se não existir, criar um novo
try:
    wb = load_workbook("./excel_tempos/tempos.xlsx")
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    #ws.append(["Data", "Função", "Tempo (Seg)", "1ª consulta"])
    ws.append(["Data", "Tempo (Seg)", "1ª consulta", "Origem/Destino",  "Sobreposição", "Zona"])  # Cabeçalhos para 1ª consulta


def iniciar_tempo():
    global start_time
    start_time = time.time()

# Função para cronometrar o tempo e salvar no Excel
def pagar_e_registar(function_name):
    global start_time, primeira_consulta
    
    if start_time is not None:
        # Registrar o tempo final
        end_time = time.time()

        # Calcular o tempo decorrido
        elapsed_time = end_time - start_time

        # Determinar se é a primeira consulta ou não
        if primeira_consulta:
            consulta_status = "sim"
        else:
            consulta_status = "não"
        
        # Obter os dados da função getInputs()
        origem_destino, sobreponivel, zona = getInputs()
        

        
        # Encontrar a próxima linha vazia
        next_row = ws.max_row + 1

        # Adicionar data, nome da função, tempo decorrido e status da primeira consulta
        ws.cell(row=next_row, column=1, value=datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        #ws.cell(row=next_row, column=2, value=function_name)
        ws.cell(row=next_row, column=2, value=elapsed_time)
        ws.cell(row=next_row, column=3, value=consulta_status)  # Exporta para o Excel se é a 1ª consulta ou não (Sim/não)
        ws.cell(row=next_row, column=4, value=origem_destino)
        ws.cell(row=next_row, column=5, value=sobreponivel)
        ws.cell(row=next_row, column=6, value=zona)

        # Guardar o ficheiro Excel
        wb.save("./excel_tempos/tempos.xlsx")
        
        # Registar que deixa de ser a primeira consulta, após guardar no Excel
        primeira_consulta = False
        start_time = time.time()  # Reiniciar a cronometragem
